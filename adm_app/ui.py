from __future__ import annotations

from datetime import datetime
import os
from pathlib import Path
import re
import shutil
import sqlite3
import subprocess
import sys
import threading
import zipfile
from openpyxl import Workbook, load_workbook

from PySide6.QtCore import QEasingCurve, QPointF, Property, QPropertyAnimation, QRect, QRectF, QSize, QTimer, Qt
from PySide6.QtGui import QColor, QIcon, QPainter, QPalette
from PySide6.QtWidgets import (
    QApplication,
    QDialog,
    QFileDialog,
    QComboBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMainWindow,
    QMessageBox,
    QInputDialog,
    QPushButton,
    QProgressBar,
    QSizePolicy,
    QSplitter,
    QStackedWidget,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QTreeWidget,
    QTreeWidgetItem,
    QVBoxLayout,
    QWidget,
    QAbstractItemView,
    QCheckBox,
)

from . import __version__
from .excel_parser import detect_header
from .i18n import normalize_language, tr
from .indexer import parse_document_part_and_revision, run_index
from .mapping import parse_qty
from .settings_store import AppSettings, save_settings
from .search import (
    get_article,
    get_article_by_number,
    get_article_bom_lines,
    get_article_ids_by_numbers,
    get_documents_for_part_revision,
    get_documents_for_link,
    get_part_detail,
    get_part_usages,
    get_unlinked_documents,
    list_articles,
)
from .update_check import fetch_latest_github_release, is_newer_version

RELEASES_URL = "https://github.com/HanPet-96/ADM-Armon-Data-management/releases"

try:
    from PySide6.QtPdf import QPdfDocument
    from PySide6.QtPdfWidgets import QPdfView

    PDF_PREVIEW_AVAILABLE = True
except Exception:
    QPdfDocument = None  # type: ignore[assignment]
    QPdfView = None  # type: ignore[assignment]
    PDF_PREVIEW_AVAILABLE = False


def resolve_app_icon_path() -> Path | None:
    candidates = [
        Path(sys.executable).resolve().parent / "Icon.ico",
        Path(__file__).resolve().parents[1] / "Icon.ico",
        Path.cwd() / "Icon.ico",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def get_app_icon() -> QIcon | None:
    icon_path = resolve_app_icon_path()
    if icon_path is None:
        return None
    icon = QIcon(str(icon_path))
    if icon.isNull():
        return None
    return icon


class PartDialog(QDialog):
    def __init__(
        self,
        conn: sqlite3.Connection,
        part_number: str,
        language: str = "en",
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self.conn = conn
        self.language = normalize_language(language)
        self.part = get_part_detail(conn, part_number)
        self.setWindowTitle(tr(self.language, "part_title", part_number=part_number))
        icon = get_app_icon()
        if icon is not None:
            self.setWindowIcon(icon)
        self.resize(760, 420)
        layout = QVBoxLayout(self)

        title = QLabel(
            f"<b>{self.part['part_number'] if self.part else part_number}</b> - {self.part['description'] if self.part and self.part['description'] else ''}"
        )
        layout.addWidget(title)
        layout.addWidget(
            QLabel(
                tr(
                    self.language,
                    "part_type",
                    part_type=self.part["part_type"] if self.part and self.part["part_type"] else "-",
                )
            )
        )

        self.usage_table = QTableWidget(0, 5)
        self.usage_table.setHorizontalHeaderLabels(["Article", "Title", "Qty", "Revision", "Material"])
        self.usage_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.usage_table.horizontalHeader().setStretchLastSection(True)
        self.usage_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        layout.addWidget(self.usage_table)

        self.docs_list = QListWidget()
        layout.addWidget(QLabel(tr(self.language, "part_docs")))
        layout.addWidget(self.docs_list)

        if self.part:
            usages = get_part_usages(conn, int(self.part["id"]))
            self.usage_table.setRowCount(len(usages))
            for row_idx, row in enumerate(usages):
                values = [
                    row["article_number"],
                    row["title"] or "",
                    "" if row["qty"] is None else str(row["qty"]),
                    row["revision"] or "",
                    row["material"] or "",
                ]
                for col_idx, value in enumerate(values):
                    self.usage_table.setItem(row_idx, col_idx, QTableWidgetItem(value))

            docs = get_documents_for_link(conn, "part", int(self.part["id"]))
            for doc in docs:
                label = doc["filename"]
                if doc["part_revision"]:
                    label = f"{label} (rev {doc['part_revision']})"
                item = QListWidgetItem(label)
                item.setData(Qt.ItemDataRole.UserRole, doc["path"])
                self.docs_list.addItem(item)

        self.docs_list.itemDoubleClicked.connect(lambda item: open_file(item.data(Qt.ItemDataRole.UserRole)))


class UnlinkedDocsDialog(QDialog):
    def __init__(self, conn: sqlite3.Connection, language: str = "en", parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.language = normalize_language(language)
        self.setWindowTitle(tr(self.language, "unlinked_title"))
        icon = get_app_icon()
        if icon is not None:
            self.setWindowIcon(icon)
        self.resize(980, 520)
        layout = QVBoxLayout(self)
        self.table = QTableWidget(0, 3)
        self.table.setHorizontalHeaderLabels(
            [
                tr(self.language, "unlinked_col_filename"),
                tr(self.language, "unlinked_col_reason"),
                tr(self.language, "unlinked_col_path"),
            ]
        )
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        layout.addWidget(self.table)
        rows = get_unlinked_documents(conn)
        self.table.setRowCount(len(rows))
        for row_idx, row in enumerate(rows):
            self.table.setItem(row_idx, 0, QTableWidgetItem(row["filename"] or ""))
            self.table.setItem(row_idx, 1, QTableWidgetItem(row["link_reason"] or ""))
            path_item = QTableWidgetItem(row["path"] or "")
            path_item.setData(Qt.ItemDataRole.UserRole, row["path"])
            self.table.setItem(row_idx, 2, path_item)
        self.table.itemDoubleClicked.connect(self.open_selected_path)

    def open_selected_path(self, item: QTableWidgetItem) -> None:
        path_item = self.table.item(item.row(), 2)
        if path_item:
            open_file(str(path_item.data(Qt.ItemDataRole.UserRole)))


class RevisionSuggestionDialog(QDialog):
    def __init__(self, rows: list[dict[str, object]], language: str = "en", parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.language = normalize_language(language)
        self.setWindowTitle(tr(self.language, "revision_suggest_title"))
        self.resize(900, 420)
        layout = QVBoxLayout(self)
        self.table = QTableWidget(0, 5)
        self.table.setHorizontalHeaderLabels(
            [
                tr(self.language, "revision_col_apply"),
                tr(self.language, "revision_col_item"),
                tr(self.language, "revision_col_part"),
                tr(self.language, "revision_col_current"),
                tr(self.language, "revision_col_found"),
            ]
        )
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        layout.addWidget(self.table)

        self.table.setRowCount(len(rows))
        for idx, row in enumerate(rows):
            apply_item = QTableWidgetItem("")
            apply_item.setFlags(apply_item.flags() | Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled)
            apply_item.setCheckState(Qt.CheckState.Checked)
            apply_item.setData(Qt.ItemDataRole.UserRole, row)
            self.table.setItem(idx, 0, apply_item)
            self.table.setItem(idx, 1, QTableWidgetItem(str(row.get("item_no") or "")))
            self.table.setItem(idx, 2, QTableWidgetItem(str(row.get("part_number") or "")))
            self.table.setItem(idx, 3, QTableWidgetItem(str(row.get("current_revision") or "")))
            self.table.setItem(idx, 4, QTableWidgetItem(str(row.get("suggested_revision") or "")))

        actions = QHBoxLayout()
        self.apply_button = QPushButton(tr(self.language, "settings_save"))
        self.cancel_button = QPushButton(tr(self.language, "settings_cancel"))
        self.apply_button.clicked.connect(self.accept)
        self.cancel_button.clicked.connect(self.reject)
        actions.addStretch(1)
        actions.addWidget(self.apply_button)
        actions.addWidget(self.cancel_button)
        layout.addLayout(actions)

    def selected_rows(self) -> list[dict[str, object]]:
        selected: list[dict[str, object]] = []
        for row_idx in range(self.table.rowCount()):
            apply_item = self.table.item(row_idx, 0)
            if apply_item is None or apply_item.checkState() != Qt.CheckState.Checked:
                continue
            payload = apply_item.data(Qt.ItemDataRole.UserRole)
            if isinstance(payload, dict):
                selected.append(payload)
        return selected


class HelpDialog(QDialog):
    def __init__(self, language: str = "en", parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.language = normalize_language(language)
        self.setWindowTitle(tr(self.language, "help_dialog_title"))
        icon = get_app_icon()
        if icon is not None:
            self.setWindowIcon(icon)
        self.resize(860, 620)
        layout = QVBoxLayout(self)

        content = QTextEdit()
        content.setReadOnly(True)
        content.setMarkdown(tr(self.language, "help_dialog_markdown"))
        layout.addWidget(content)

        actions = QHBoxLayout()
        close_button = QPushButton(tr(self.language, "btn_close"))
        close_button.clicked.connect(self.accept)
        actions.addStretch(1)
        actions.addWidget(close_button)
        layout.addLayout(actions)


class ToggleSwitch(QCheckBox):
    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self._offset = 0.0
        self._anim = QPropertyAnimation(self, b"offset", self)
        self._anim.setDuration(140)
        self._anim.setEasingCurve(QEasingCurve.Type.InOutCubic)
        self.stateChanged.connect(self._animate_to_state)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setFixedSize(52, 28)

    def sizeHint(self) -> QSize:
        return QSize(52, 28)

    def hitButton(self, pos) -> bool:  # type: ignore[override]
        return self.contentsRect().contains(pos)

    def _animate_to_state(self) -> None:
        self._anim.stop()
        self._anim.setStartValue(self._offset)
        self._anim.setEndValue(1.0 if self.isChecked() else 0.0)
        self._anim.start()

    def get_offset(self) -> float:
        return self._offset

    def set_offset(self, value: float) -> None:
        self._offset = float(value)
        self.update()

    offset = Property(float, get_offset, set_offset)

    def paintEvent(self, _event) -> None:  # type: ignore[override]
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing, True)

        rect = QRectF(1, 1, self.width() - 2, self.height() - 2)
        track_on = QColor("#97C21E")
        track_off = QColor(130, 130, 130)
        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(track_on if self.isChecked() else track_off)
        painter.drawRoundedRect(rect, rect.height() / 2, rect.height() / 2)

        margin = 3.0
        knob_d = rect.height() - margin * 2
        x_min = rect.x() + margin
        x_max = rect.right() - margin - knob_d
        knob_x = x_min + (x_max - x_min) * self._offset
        knob_rect = QRectF(knob_x, rect.y() + margin, knob_d, knob_d)
        painter.setBrush(QColor(255, 255, 255))
        painter.drawEllipse(knob_rect)


class BackdropWidget(QWidget):
    def __init__(self, on_click, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.on_click = on_click
        self._opacity = 0.0

    def get_opacity(self) -> float:
        return self._opacity

    def set_opacity(self, value: float) -> None:
        self._opacity = max(0.0, min(1.0, float(value)))
        self.update()

    opacity = Property(float, get_opacity, set_opacity)

    def mousePressEvent(self, event) -> None:  # type: ignore[override]
        if event.button() == Qt.MouseButton.LeftButton:
            self.on_click()
            event.accept()
            return
        super().mousePressEvent(event)

    def paintEvent(self, _event) -> None:  # type: ignore[override]
        if self._opacity <= 0.0:
            return
        painter = QPainter(self)
        alpha = int(170 * self._opacity)
        painter.fillRect(self.rect(), QColor(0, 0, 0, alpha))


class SettingsDialog(QDialog):
    def __init__(
        self,
        data_root: str,
        theme_mode: str,
        language: str,
        developer_mode: bool,
        developer_toggle_available: bool,
        parent: QWidget | None = None,
    ) -> None:
        super().__init__(parent)
        self.language = normalize_language(language)
        self.developer_toggle_available = developer_toggle_available
        self.setWindowTitle(tr(self.language, "settings_title"))
        icon = get_app_icon()
        if icon is not None:
            self.setWindowIcon(icon)
        self.resize(700, 280)
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(tr(self.language, "settings_datastruct")))
        row = QHBoxLayout()
        self.data_root_input = QLineEdit(data_root)
        browse_button = QPushButton(tr(self.language, "settings_browse"))
        browse_button.clicked.connect(self.browse_data_root)
        row.addWidget(self.data_root_input)
        row.addWidget(browse_button)
        layout.addLayout(row)

        theme_row = QHBoxLayout()
        theme_row.addWidget(QLabel(tr(self.language, "settings_dark_mode")))
        self.theme_toggle = ToggleSwitch()
        self.theme_toggle.setChecked(theme_mode.lower() == "dark")
        theme_row.addWidget(self.theme_toggle)
        theme_row.addStretch(1)
        layout.addLayout(theme_row)

        language_row = QHBoxLayout()
        language_row.addWidget(QLabel(tr(self.language, "settings_language")))
        self.language_combo = QComboBox()
        self.language_combo.addItem(tr(self.language, "lang_name_en"), "en")
        self.language_combo.addItem(tr(self.language, "lang_name_nl"), "nl")
        current_idx = 1 if self.language == "nl" else 0
        self.language_combo.setCurrentIndex(current_idx)
        language_row.addWidget(self.language_combo)
        language_row.addStretch(1)
        layout.addLayout(language_row)

        self.developer_toggle: ToggleSwitch | None = None
        if self.developer_toggle_available:
            dev_row = QHBoxLayout()
            dev_row.addWidget(QLabel(tr(self.language, "settings_developer_mode")))
            self.developer_toggle = ToggleSwitch()
            self.developer_toggle.setChecked(bool(developer_mode))
            dev_row.addWidget(self.developer_toggle)
            dev_row.addStretch(1)
            layout.addLayout(dev_row)

        actions = QHBoxLayout()
        self.check_updates_button = QPushButton(tr(self.language, "settings_check_updates"))
        save_button = QPushButton(tr(self.language, "settings_save"))
        cancel_button = QPushButton(tr(self.language, "settings_cancel"))
        self.check_updates_button.clicked.connect(self.check_for_updates_clicked)
        save_button.clicked.connect(self.accept)
        cancel_button.clicked.connect(self.reject)
        actions.addWidget(self.check_updates_button)
        actions.addStretch(1)
        actions.addWidget(save_button)
        actions.addWidget(cancel_button)
        layout.addLayout(actions)

    def browse_data_root(self) -> None:
        selected = QFileDialog.getExistingDirectory(
            self,
            tr(self.language, "settings_datastruct"),
            self.data_root_input.text(),
        )
        if selected:
            self.data_root_input.setText(selected)

    def selected_data_root(self) -> str:
        return self.data_root_input.text().strip()

    def selected_theme_mode(self) -> str:
        return "dark" if self.theme_toggle.isChecked() else "light"

    def selected_language(self) -> str:
        return normalize_language(str(self.language_combo.currentData()))

    def selected_developer_mode(self) -> bool:
        if self.developer_toggle is None:
            return False
        return bool(self.developer_toggle.isChecked())

    def check_for_updates_clicked(self) -> None:
        latest = fetch_latest_github_release()
        if not latest:
            QMessageBox.information(
                self,
                tr(self.language, "update_check_title"),
                tr(self.language, "update_check_failed"),
            )
            return
        latest_tag = str(latest.get("tag_name") or "").strip()
        if not latest_tag:
            QMessageBox.information(
                self,
                tr(self.language, "update_check_title"),
                tr(self.language, "update_check_failed"),
            )
            return
        if is_newer_version(__version__, latest_tag):
            answer = QMessageBox.question(
                self,
                tr(self.language, "update_available_title"),
                tr(self.language, "update_available_msg", installed=__version__, latest=latest_tag),
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.Yes,
            )
            if answer == QMessageBox.StandardButton.Yes:
                open_file(RELEASES_URL)
            return
        QMessageBox.information(
            self,
            tr(self.language, "up_to_date_title"),
            tr(self.language, "up_to_date_msg", version=__version__),
        )


class MainWindow(QMainWindow):
    def __init__(
        self,
        conn: sqlite3.Connection,
        data_root: str,
        theme_mode: str = "light",
        has_seen_help: bool = False,
        language: str = "en",
        developer_mode: bool = False,
    ) -> None:
        super().__init__()
        self.conn = conn
        self.data_root = data_root
        self.theme_mode = theme_mode if theme_mode in {"light", "dark"} else "light"
        self.has_seen_help = has_seen_help
        self.language = normalize_language(language)
        self.developer_mode = bool(developer_mode)
        self.developer_toggle_available = os.getenv("COMPUTERNAME", "").strip().lower() in {
            "laptop-han",
            "rendlaptop",
        }
        if not self.developer_toggle_available:
            self.developer_mode = False
        self._update_check_done = False
        self.active_search_term = ""
        self.current_article_id: int | None = None
        self.current_article_source_bom_path = ""
        self._bom_dirty = False
        self._suspend_bom_item_changed = False
        self._pending_bom_edits: dict[int, dict[str, object]] = {}
        self.current_pdf_path = ""
        self.current_pdf_page = 0
        self.current_pdf_page_count = 0
        self.order_cart: dict[str, dict[str, object]] = {}
        self.order_drawer_open = False
        self.order_drawer_width = 380
        self.order_drawer_anim: QPropertyAnimation | None = None
        self.order_backdrop_anim: QPropertyAnimation | None = None
        self.setWindowTitle(tr(self.language, "window_title", version=__version__))
        icon = get_app_icon()
        if icon is not None:
            self.setWindowIcon(icon)
        self.resize(1250, 740)

        root = QWidget()
        self.setCentralWidget(root)
        root_layout = QVBoxLayout(root)

        top_actions = QHBoxLayout()
        self.reindex_button = QPushButton(tr(self.language, "btn_reindex"))
        self.unlinked_button = QPushButton(tr(self.language, "btn_unlinked"))
        self.open_order_button = QPushButton(tr(self.language, "btn_order_list"))
        self.settings_button = QPushButton(tr(self.language, "btn_settings"))
        self.help_button = QPushButton(tr(self.language, "btn_help"))
        self.help_button.setToolTip(tr(self.language, "tooltip_help"))
        top_actions.addStretch(1)
        top_actions.addWidget(self.reindex_button)
        top_actions.addWidget(self.unlinked_button)
        top_actions.addWidget(self.open_order_button)
        top_actions.addWidget(self.settings_button)
        top_actions.addWidget(self.help_button)
        root_layout.addLayout(top_actions)

        splitter = QSplitter(Qt.Orientation.Horizontal)
        root_layout.addWidget(splitter)

        left_area = QWidget()
        left_area_layout = QVBoxLayout(left_area)
        left_area_layout.setContentsMargins(0, 0, 0, 0)
        left_controls = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText(tr(self.language, "search_placeholder"))
        self.search_button = QPushButton(tr(self.language, "btn_search"))
        left_controls.addWidget(self.search_input, 1)
        left_controls.addWidget(self.search_button)
        left_area_layout.addLayout(left_controls)
        splitter.addWidget(left_area)

        self.article_table = QTableWidget(0, 2)
        self.article_table.setHorizontalHeaderLabels(
            [
                tr(self.language, "article_col_article"),
                tr(self.language, "article_col_title"),
            ]
        )
        self.article_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.article_table.horizontalHeader().setStretchLastSection(True)
        self.article_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        left_area_layout.addWidget(self.article_table)

        right = QWidget()
        right_layout = QVBoxLayout(right)
        self.article_title_label = QLabel(tr(self.language, "lbl_select_article"))
        right_layout.addWidget(self.article_title_label)

        self.right_content_splitter = QSplitter(Qt.Orientation.Vertical)
        right_layout.addWidget(self.right_content_splitter)

        bom_container = QWidget()
        bom_layout = QVBoxLayout(bom_container)
        bom_layout.setContentsMargins(0, 0, 0, 0)
        bom_controls = QHBoxLayout()
        self.expand_all_button = QPushButton(tr(self.language, "btn_expand_all"))
        self.collapse_all_button = QPushButton(tr(self.language, "btn_collapse_all"))
        self.add_to_order_button = QPushButton(tr(self.language, "btn_add_to_order"))
        self.revision_suggest_button = QPushButton(tr(self.language, "btn_revision_check"))
        self.save_bom_button = QPushButton(tr(self.language, "btn_save_bom"))
        self.revision_suggest_button.setVisible(self.developer_mode)
        self.save_bom_button.setVisible(False)
        bom_controls.addWidget(self.expand_all_button)
        bom_controls.addWidget(self.collapse_all_button)
        bom_controls.addWidget(self.add_to_order_button)
        bom_controls.addWidget(self.revision_suggest_button)
        bom_controls.addWidget(self.save_bom_button)
        bom_controls.addStretch(1)
        bom_layout.addLayout(bom_controls)
        self.bom_tree = QTreeWidget()
        self.bom_tree.setColumnCount(9)
        self.bom_tree.setHeaderLabels(
            [
                tr(self.language, "bom_col_item"),
                tr(self.language, "bom_col_part"),
                tr(self.language, "bom_col_rev"),
                tr(self.language, "bom_col_desc"),
                tr(self.language, "bom_col_material"),
                tr(self.language, "bom_col_finish"),
                tr(self.language, "bom_col_qty"),
                tr(self.language, "bom_col_type"),
                tr(self.language, "bom_col_status"),
            ]
        )
        self.bom_tree.header().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.bom_tree.header().setStretchLastSection(True)
        self.bom_tree.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        if self.developer_mode:
            self.bom_tree.setEditTriggers(
                QAbstractItemView.EditTrigger.DoubleClicked
                | QAbstractItemView.EditTrigger.SelectedClicked
                | QAbstractItemView.EditTrigger.EditKeyPressed
            )
        else:
            self.bom_tree.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        bom_layout.addWidget(self.bom_tree)
        self.right_content_splitter.addWidget(bom_container)

        docs_area = QWidget()
        docs_area_layout = QVBoxLayout(docs_area)
        docs_area_layout.setContentsMargins(0, 0, 0, 0)
        self.docs_context_label = QLabel(tr(self.language, "lbl_docs_article"))
        self.docs_context_label.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Fixed)
        docs_area_layout.addWidget(self.docs_context_label)

        docs_preview_splitter = QSplitter(Qt.Orientation.Horizontal)
        docs_area_layout.addWidget(docs_preview_splitter)
        docs_area_layout.setStretch(1, 1)
        self.right_content_splitter.addWidget(docs_area)
        self.right_content_splitter.setStretchFactor(0, 1)
        self.right_content_splitter.setStretchFactor(1, 1)

        docs_list_container = QWidget()
        docs_list_layout = QVBoxLayout(docs_list_container)
        docs_list_layout.setContentsMargins(0, 0, 0, 0)
        self.docs_list = QListWidget()
        docs_list_layout.addWidget(self.docs_list)
        docs_preview_splitter.addWidget(docs_list_container)

        preview_container = QWidget()
        preview_layout = QVBoxLayout(preview_container)
        preview_layout.setContentsMargins(0, 0, 0, 0)
        preview_title = QLabel(tr(self.language, "lbl_pdf_preview"))
        preview_title.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Fixed)
        preview_layout.addWidget(preview_title)
        self.preview_controls_row = QHBoxLayout()
        self.preview_prev_button = QPushButton("<")
        self.preview_next_button = QPushButton(">")
        self.preview_page_label = QLabel("-")
        self.preview_page_hint = QLabel("")
        self.preview_prev_button.clicked.connect(self.preview_prev_page)
        self.preview_next_button.clicked.connect(self.preview_next_page)
        self.preview_controls_row.addWidget(self.preview_prev_button)
        self.preview_controls_row.addWidget(self.preview_next_button)
        self.preview_controls_row.addWidget(self.preview_page_label)
        self.preview_controls_row.addStretch(1)
        self.preview_controls_row.addWidget(self.preview_page_hint)
        preview_layout.addLayout(self.preview_controls_row)
        self.preview_stack = QStackedWidget()
        self.preview_message = QLabel(tr(self.language, "lbl_preview_default"))
        self.preview_message.setWordWrap(True)
        self.preview_message.setAlignment(Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignLeft)
        self.preview_message.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.preview_stack.addWidget(self.preview_message)
        self.pdf_document = None
        self.pdf_view = None
        if PDF_PREVIEW_AVAILABLE:
            self.pdf_document = QPdfDocument(self)
            self.pdf_view = QPdfView()
            self.pdf_view.setDocument(self.pdf_document)
            self.pdf_view.setPageMode(QPdfView.PageMode.SinglePage)
            self.pdf_view.setZoomMode(QPdfView.ZoomMode.FitInView)
            self.pdf_view.verticalScrollBar().setValue(0)
            self.pdf_view.horizontalScrollBar().setValue(0)
            self.preview_stack.addWidget(self.pdf_view)
        preview_layout.addWidget(self.preview_stack)
        docs_preview_splitter.addWidget(preview_container)
        docs_preview_splitter.setStretchFactor(0, 2)
        docs_preview_splitter.setStretchFactor(1, 3)

        splitter.addWidget(right)
        splitter.setStretchFactor(0, 2)
        splitter.setStretchFactor(1, 3)

        self.order_backdrop = BackdropWidget(self.close_order_drawer, root)
        self.order_backdrop.setObjectName("orderBackdrop")
        self.order_backdrop.set_opacity(0.0)
        self.order_backdrop.hide()

        self.order_drawer = QWidget(root)
        self.order_drawer.setObjectName("orderDrawer")
        self.order_drawer.setStyleSheet("#orderDrawer { border: 1px solid #D1D5DB; background: palette(base); }")
        drawer_layout = QVBoxLayout(self.order_drawer)
        drawer_layout.setContentsMargins(10, 10, 10, 10)
        drawer_head = QHBoxLayout()
        drawer_title = QLabel(tr(self.language, "order_title"))
        self.close_order_drawer_button = QPushButton(tr(self.language, "btn_close"))
        drawer_head.addWidget(drawer_title)
        drawer_head.addStretch(1)
        drawer_head.addWidget(self.close_order_drawer_button)
        drawer_layout.addLayout(drawer_head)
        self.order_table = QTableWidget(0, 3)
        self.order_table.setColumnCount(4)
        self.order_table.setHorizontalHeaderLabels(
            [
                tr(self.language, "order_tbl_part"),
                tr(self.language, "order_tbl_rev"),
                tr(self.language, "order_tbl_qty"),
                "",
            ]
        )
        self.order_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.order_table.horizontalHeader().setStretchLastSection(True)
        self.order_table.setColumnWidth(3, 44)
        self.order_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        drawer_layout.addWidget(self.order_table)
        drawer_actions = QHBoxLayout()
        self.export_order_button = QPushButton(tr(self.language, "btn_export_xlsx_zip"))
        self.clear_order_button = QPushButton(tr(self.language, "btn_clear"))
        drawer_actions.addWidget(self.export_order_button)
        drawer_actions.addWidget(self.clear_order_button)
        drawer_layout.addLayout(drawer_actions)

        self.search_button.clicked.connect(self.refresh_articles)
        self.search_input.returnPressed.connect(self.refresh_articles)
        self.reindex_button.clicked.connect(self.reindex)
        self.unlinked_button.clicked.connect(self.open_unlinked_docs)
        self.add_to_order_button.clicked.connect(self.add_selected_bom_to_order)
        self.revision_suggest_button.clicked.connect(self.run_revision_suggestions)
        self.save_bom_button.clicked.connect(self.save_bom_edits)
        self.open_order_button.clicked.connect(self.toggle_order_drawer)
        self.close_order_drawer_button.clicked.connect(self.close_order_drawer)
        self.export_order_button.clicked.connect(self.export_order_bundle)
        self.clear_order_button.clicked.connect(self.clear_order_cart)
        self.settings_button.clicked.connect(self.open_settings)
        self.help_button.clicked.connect(self.open_help_manual_clicked)
        self.article_table.itemSelectionChanged.connect(self.load_selected_article)
        self.docs_list.itemDoubleClicked.connect(lambda item: open_file(item.data(Qt.ItemDataRole.UserRole)))
        self.docs_list.itemSelectionChanged.connect(self.preview_selected_document)
        self.bom_tree.itemDoubleClicked.connect(self.open_bom_line)
        self.bom_tree.itemSelectionChanged.connect(self.load_docs_for_selected_bom_line)
        self.bom_tree.itemChanged.connect(self.on_bom_item_changed)
        self.expand_all_button.clicked.connect(self.expand_bom_children_only)
        self.collapse_all_button.clicked.connect(self.collapse_bom_children_only)

        self.apply_translations()
        self.refresh_articles()
        QTimer.singleShot(0, self._set_initial_split_sizes)
        QTimer.singleShot(0, self.layout_order_drawer)
        self.apply_theme(self.theme_mode)
        if not self.has_seen_help:
            QTimer.singleShot(250, self.open_help_manual_first_run)
        QTimer.singleShot(900, self.check_for_updates)

    def apply_translations(self) -> None:
        self.setWindowTitle(tr(self.language, "window_title", version=__version__))
        self.reindex_button.setText(tr(self.language, "btn_reindex"))
        self.unlinked_button.setText(tr(self.language, "btn_unlinked"))
        self.open_order_button.setText(tr(self.language, "btn_order_list"))
        self.settings_button.setText(tr(self.language, "btn_settings"))
        self.help_button.setText(tr(self.language, "btn_help"))
        self.help_button.setToolTip(tr(self.language, "tooltip_help"))
        self.search_input.setPlaceholderText(tr(self.language, "search_placeholder"))
        self.search_button.setText(tr(self.language, "btn_search"))
        self.article_title_label.setText(tr(self.language, "lbl_select_article"))
        self.expand_all_button.setText(tr(self.language, "btn_expand_all"))
        self.collapse_all_button.setText(tr(self.language, "btn_collapse_all"))
        self.add_to_order_button.setText(tr(self.language, "btn_add_to_order"))
        self.revision_suggest_button.setText(tr(self.language, "btn_revision_check"))
        self.save_bom_button.setText(tr(self.language, "btn_save_bom"))
        self.docs_context_label.setText(tr(self.language, "lbl_docs_article"))
        self.preview_message.setText(tr(self.language, "lbl_preview_default"))
        self.preview_page_hint.setText("")
        self.close_order_drawer_button.setText(tr(self.language, "btn_close"))
        self.export_order_button.setText(tr(self.language, "btn_export_xlsx_zip"))
        self.clear_order_button.setText(tr(self.language, "btn_clear"))
        self.article_table.setHorizontalHeaderLabels(
            [
                tr(self.language, "article_col_article"),
                tr(self.language, "article_col_title"),
            ]
        )
        self.bom_tree.setHeaderLabels(
            [
                tr(self.language, "bom_col_item"),
                tr(self.language, "bom_col_part"),
                tr(self.language, "bom_col_rev"),
                tr(self.language, "bom_col_desc"),
                tr(self.language, "bom_col_material"),
                tr(self.language, "bom_col_finish"),
                tr(self.language, "bom_col_qty"),
                tr(self.language, "bom_col_type"),
                tr(self.language, "bom_col_status"),
            ]
        )
        self.order_table.setHorizontalHeaderLabels(
            [
                tr(self.language, "order_tbl_part"),
                tr(self.language, "order_tbl_rev"),
                tr(self.language, "order_tbl_qty"),
                "",
            ]
        )
        self.update_preview_controls()

    def refresh_articles(self) -> None:
        self.active_search_term = self.search_input.text().strip()
        rows = list_articles(self.conn, self.active_search_term)
        self.article_table.setRowCount(len(rows))
        for row_idx, row in enumerate(rows):
            article_item = QTableWidgetItem(row["article_number"])
            article_item.setData(Qt.ItemDataRole.UserRole, int(row["id"]))
            self.article_table.setItem(row_idx, 0, article_item)
            self.article_table.setItem(row_idx, 1, QTableWidgetItem(row["title"] or ""))
        if rows:
            self.article_table.selectRow(0)

    def load_selected_article(self) -> None:
        row = self.article_table.currentRow()
        if row < 0:
            return
        article_id_item = self.article_table.item(row, 0)
        if article_id_item is None:
            return
        article_id = article_id_item.data(Qt.ItemDataRole.UserRole)
        if article_id is None:
            return
        self.display_article_by_id(int(article_id))

    def select_article_in_table(self, article_id: int) -> bool:
        for row in range(self.article_table.rowCount()):
            article_id_item = self.article_table.item(row, 0)
            if article_id_item is None:
                continue
            row_id = article_id_item.data(Qt.ItemDataRole.UserRole)
            if row_id is None:
                continue
            if int(row_id) == article_id:
                self.article_table.selectRow(row)
                self.article_table.scrollToItem(article_id_item)
                return True
        return False

    def display_article_by_id(self, article_id: int) -> None:
        article = get_article(self.conn, article_id)
        if not article:
            return
        self._suspend_bom_item_changed = True
        self.current_article_id = article_id
        self.current_article_source_bom_path = str(article["source_bom_path"] or "")
        self._bom_dirty = False
        self._pending_bom_edits = {}
        self.save_bom_button.setVisible(False)
        self.article_title_label.setText(
            f"<b>Article {article['article_number']}</b> - {article['title'] or ''}<br>{article['source_bom_filename'] or ''}"
        )
        lines = get_article_bom_lines(self.conn, article_id)
        article_ref_map = self.resolve_article_ref_map(lines)
        self.bom_tree.clear()
        root_item = QTreeWidgetItem(
            [
                tr(self.language, "root_label"),
                str(article["article_number"] or ""),
                "",
                str(article["title"] or ""),
                "",
                "",
                "1",
                tr(self.language, "root_type"),
                "",
            ]
        )
        root_item.setData(0, Qt.ItemDataRole.UserRole, int(article_id))
        root_item.setData(1, Qt.ItemDataRole.UserRole, str(article["article_number"] or ""))
        root_item.setData(5, Qt.ItemDataRole.UserRole, "article")
        self.bom_tree.addTopLevelItem(root_item)
        items_by_item_no: dict[str, QTreeWidgetItem] = {}
        all_items: list[QTreeWidgetItem] = []
        first_item: QTreeWidgetItem | None = root_item
        for line in lines:
            item_no = str(line["item_no"] or "").strip()
            values = [
                item_no,
                line["part_number"] or "",
                line["revision"] or "",
                line["description"] or "",
                line["material"] or "",
                line["finish"] or "",
                "" if line["qty"] is None else str(line["qty"]),
                line["line_type"] or "",
                line["status"] or "",
            ]
            tree_item = QTreeWidgetItem(values)
            tree_item.setData(0, Qt.ItemDataRole.UserRole, line["id"])
            tree_item.setData(1, Qt.ItemDataRole.UserRole, line["part_number"] or "")
            tree_item.setData(2, Qt.ItemDataRole.UserRole, line["revision"] or "")
            tree_item.setData(3, Qt.ItemDataRole.UserRole, line["part_id"])
            tree_item.setData(4, Qt.ItemDataRole.UserRole, float(line["qty"]) if line["qty"] is not None else 1.0)
            tree_item.setData(7, Qt.ItemDataRole.UserRole, str(line["source_sheet"] or ""))
            tree_item.setData(8, Qt.ItemDataRole.UserRole, int(line["source_row_number"] or 0))
            child_article_id = article_ref_map.get(str(line["part_number"] or "").strip())
            if child_article_id is not None:
                tree_item.setData(5, Qt.ItemDataRole.UserRole, "assembly_ref")
                tree_item.setData(6, Qt.ItemDataRole.UserRole, int(child_article_id))
            else:
                tree_item.setData(5, Qt.ItemDataRole.UserRole, "part")
                tree_item.setData(6, Qt.ItemDataRole.UserRole, None)
            if self.developer_mode:
                tree_item.setFlags(tree_item.flags() | Qt.ItemFlag.ItemIsEditable)
            self.apply_status_style(tree_item, values[8], 8)

            parent_item = items_by_item_no.get(self.parent_item_no(item_no))
            if parent_item is not None:
                parent_item.addChild(tree_item)
            else:
                root_item.addChild(tree_item)
            if item_no:
                items_by_item_no[item_no] = tree_item
            all_items.append(tree_item)

        self.bom_tree.collapseAll()
        root_item.setExpanded(True)
        self.load_article_documents(article_id)
        selected_item = self.find_bom_item_for_search(all_items, self.active_search_term) or first_item
        if selected_item is not None:
            self.expand_path_to_item(selected_item)
            self.bom_tree.setCurrentItem(selected_item)
            self.bom_tree.scrollToItem(selected_item)
        self._suspend_bom_item_changed = False

    def parent_item_no(self, item_no: str) -> str:
        value = (item_no or "").strip()
        if not value or "." not in value:
            return ""
        return value.rsplit(".", 1)[0]

    def article_number_candidates(self, part_number: str) -> list[str]:
        raw = str(part_number or "").strip()
        if not raw:
            return []
        candidates: list[str] = [raw]
        digit_tokens = re.findall(r"\d{3,}", raw)
        for token in sorted(set(digit_tokens), key=lambda v: (-len(v), v)):
            if token not in candidates:
                candidates.append(token)
            trimmed = token.lstrip("0")
            if trimmed and trimmed not in candidates:
                candidates.append(trimmed)
        return candidates

    def resolve_article_ref_map(self, lines: list[sqlite3.Row]) -> dict[str, int]:
        parts = [str(line["part_number"] or "").strip() for line in lines]
        candidates: list[str] = []
        for part in parts:
            candidates.extend(self.article_number_candidates(part))
        found = get_article_ids_by_numbers(self.conn, candidates)
        mapping: dict[str, int] = {}
        for part in parts:
            for candidate in self.article_number_candidates(part):
                article_id = found.get(candidate)
                if article_id is not None:
                    mapping[part] = int(article_id)
                    break
        return mapping

    def resolve_article_ref_id(self, part_number: str) -> int | None:
        for candidate in self.article_number_candidates(part_number):
            row = get_article_by_number(self.conn, candidate)
            if row is not None:
                return int(row["id"])
        return None

    def expand_bom_children_only(self) -> None:
        root = self.bom_tree.topLevelItem(0)
        if root is None:
            return
        for idx in range(root.childCount()):
            self._set_expanded_recursive(root.child(idx), True)

    def collapse_bom_children_only(self) -> None:
        root = self.bom_tree.topLevelItem(0)
        if root is None:
            return
        for idx in range(root.childCount()):
            self._set_expanded_recursive(root.child(idx), False)
        root.setExpanded(True)

    def _set_expanded_recursive(self, item: QTreeWidgetItem, expanded: bool) -> None:
        item.setExpanded(expanded)
        for idx in range(item.childCount()):
            self._set_expanded_recursive(item.child(idx), expanded)

    def find_bom_item_for_search(self, items: list[QTreeWidgetItem], query: str) -> QTreeWidgetItem | None:
        term = (query or "").strip().lower()
        if not term:
            return None
        exact = [item for item in items if str(item.data(1, Qt.ItemDataRole.UserRole) or "").strip().lower() == term]
        if exact:
            return exact[0]
        starts = [
            item for item in items if str(item.data(1, Qt.ItemDataRole.UserRole) or "").strip().lower().startswith(term)
        ]
        if starts:
            return starts[0]
        contains = [item for item in items if term in str(item.data(1, Qt.ItemDataRole.UserRole) or "").strip().lower()]
        if contains:
            return contains[0]
        return None

    def expand_path_to_item(self, item: QTreeWidgetItem) -> None:
        current = item.parent()
        while current is not None:
            current.setExpanded(True)
            current = current.parent()

    def _set_initial_split_sizes(self) -> None:
        total_height = self.right_content_splitter.height()
        if total_height <= 0:
            self.right_content_splitter.setSizes([1, 1])
            return
        half = max(1, total_height // 2)
        self.right_content_splitter.setSizes([half, half])

    def resizeEvent(self, event) -> None:  # type: ignore[override]
        super().resizeEvent(event)
        self.layout_order_drawer()

    def showEvent(self, event) -> None:  # type: ignore[override]
        super().showEvent(event)
        self.layout_order_drawer()

    def load_article_documents(self, article_id: int) -> None:
        self.docs_context_label.setText(tr(self.language, "lbl_docs_article"))
        self.docs_list.clear()
        docs = get_documents_for_link(self.conn, "article", article_id)
        for doc in docs:
            item = QListWidgetItem(doc["filename"])
            item.setData(Qt.ItemDataRole.UserRole, doc["path"])
            self.docs_list.addItem(item)
        self.preview_first_pdf_in_list()

    def load_docs_for_selected_bom_line(self) -> None:
        current = self.bom_tree.currentItem()
        if current is None:
            return
        node_type = str(current.data(5, Qt.ItemDataRole.UserRole) or "")
        if node_type == "article":
            article_id = int(current.data(0, Qt.ItemDataRole.UserRole))
            self.load_article_documents(article_id)
            return
        if node_type == "assembly_ref":
            child_article_id = current.data(6, Qt.ItemDataRole.UserRole)
            if child_article_id is None:
                child_article_id = self.resolve_article_ref_id(str(current.data(1, Qt.ItemDataRole.UserRole) or ""))
                if child_article_id is not None:
                    current.setData(6, Qt.ItemDataRole.UserRole, int(child_article_id))
            if child_article_id is not None:
                self.load_article_documents(int(child_article_id))
                return
        part_number = str(current.data(1, Qt.ItemDataRole.UserRole) or "").strip()
        if not part_number:
            return
        part = get_part_detail(self.conn, part_number)
        if not part:
            return
        revision = str(current.data(2, Qt.ItemDataRole.UserRole) or "").strip()
        revision = revision or None
        docs = get_documents_for_part_revision(self.conn, int(part["id"]), revision)
        self.docs_context_label.setText(
            f"{tr(self.language, 'part_docs')} ({part_number}{f', rev {revision}' if revision else ''})"
        )
        self.docs_list.clear()
        for doc in docs:
            label = doc["filename"]
            if doc["part_revision"]:
                label = f"{label} (rev {doc['part_revision']})"
            item = QListWidgetItem(label)
            item.setData(Qt.ItemDataRole.UserRole, doc["path"])
            self.docs_list.addItem(item)
        self.preview_first_pdf_in_list()

    def reindex(self) -> None:
        target_article_id = int(self.current_article_id or 0)
        stats = self.run_reindex_with_progress()
        if stats is None:
            return
        self.refresh_articles()
        if target_article_id > 0:
            if self.select_article_in_table(target_article_id):
                self.display_article_by_id(target_article_id)
            else:
                self.display_article_by_id(target_article_id)

    def run_reindex_with_progress(self) -> object | None:
        progress_dialog = QDialog(self)
        progress_dialog.setWindowTitle(tr(self.language, "reindex_progress_title"))
        progress_dialog.setModal(True)
        progress_dialog.setWindowFlag(Qt.WindowType.WindowCloseButtonHint, False)
        progress_dialog.setWindowFlag(Qt.WindowType.WindowContextHelpButtonHint, False)
        progress_dialog.setFixedSize(460, 130)
        layout = QVBoxLayout(progress_dialog)
        layout.addWidget(QLabel(tr(self.language, "reindex_progress_msg")))
        bar = QProgressBar()
        bar.setRange(0, 0)
        layout.addWidget(bar)
        status_label = QLabel(tr(self.language, "reindex_progress_running"))
        layout.addWidget(status_label)

        result: dict[str, object] = {"stats": None, "error": None, "done": False}
        data_root_path = self._path_to_root()
        db_row = self.conn.execute("PRAGMA database_list").fetchone()
        db_path = Path(str(db_row["file"] if db_row and "file" in db_row.keys() else "")).resolve()

        def _index_in_background() -> None:
            from .db import get_connection

            local_conn = get_connection(db_path)
            try:
                result["stats"] = run_index(local_conn, data_root=data_root_path)
            except Exception as exc:
                result["error"] = str(exc)
            finally:
                local_conn.close()
                result["done"] = True

        worker = threading.Thread(target=_index_in_background, daemon=True)
        worker.start()

        poll_timer = QTimer(progress_dialog)

        def _poll() -> None:
            if not bool(result.get("done")):
                return
            poll_timer.stop()
            if result.get("error"):
                status_label.setText(tr(self.language, "reindex_progress_failed"))
                progress_dialog.reject()
                return
            status_label.setText(tr(self.language, "reindex_progress_done"))
            progress_dialog.accept()

        poll_timer.setInterval(100)
        poll_timer.timeout.connect(_poll)
        poll_timer.start()
        progress_dialog.exec()
        worker.join(timeout=10.0)

        if result["error"]:
            QMessageBox.warning(
                self,
                tr(self.language, "reindex_done_title"),
                tr(self.language, "startup_reindex_error_msg", error=str(result["error"])),
            )
            return None
        return result["stats"]

    def _path_to_root(self):
        from pathlib import Path

        return Path(self.data_root).resolve()

    def open_bom_line(self, item: QTreeWidgetItem, _column: int) -> None:
        node_type = str(item.data(5, Qt.ItemDataRole.UserRole) or "")
        if node_type == "article":
            return
        if node_type == "assembly_ref":
            child_article_id = item.data(6, Qt.ItemDataRole.UserRole)
            if child_article_id is None:
                child_article_id = self.resolve_article_ref_id(str(item.data(1, Qt.ItemDataRole.UserRole) or ""))
                if child_article_id is not None:
                    item.setData(6, Qt.ItemDataRole.UserRole, int(child_article_id))
            if child_article_id is None:
                return
            if self.select_article_in_table(int(child_article_id)):
                return
            self.display_article_by_id(int(child_article_id))
            return
        if node_type == "part":
            resolved_article_id = self.resolve_article_ref_id(str(item.data(1, Qt.ItemDataRole.UserRole) or ""))
            if resolved_article_id is not None:
                item.setData(5, Qt.ItemDataRole.UserRole, "assembly_ref")
                item.setData(6, Qt.ItemDataRole.UserRole, int(resolved_article_id))
                if self.select_article_in_table(int(resolved_article_id)):
                    return
                self.display_article_by_id(int(resolved_article_id))
                return
        part_number = str(item.data(1, Qt.ItemDataRole.UserRole) or "").strip()
        if not part_number:
            return
        dlg = PartDialog(self.conn, part_number, language=self.language, parent=self)
        dlg.exec()

    def open_unlinked_docs(self) -> None:
        dlg = UnlinkedDocsDialog(self.conn, language=self.language, parent=self)
        dlg.exec()

    def add_selected_bom_to_order(self) -> None:
        current = self.bom_tree.currentItem()
        if current is None:
            QMessageBox.information(self, tr(self.language, "order_added_title"), tr(self.language, "order_select_first"))
            return
        include_dialog = QInputDialog(self)
        include_dialog.setWindowTitle(tr(self.language, "order_include_title"))
        include_dialog.setLabelText(tr(self.language, "order_include_label"))
        include_dialog.setComboBoxItems(
            [
                tr(self.language, "order_include_children"),
                tr(self.language, "order_include_selected"),
                tr(self.language, "order_include_both"),
            ]
        )
        include_dialog.setComboBoxEditable(False)
        include_dialog.resize(350, include_dialog.sizeHint().height())
        if include_dialog.exec() != QDialog.DialogCode.Accepted:
            return
        include_choice = include_dialog.textValue()
        qty_multiplier, ok = QInputDialog.getInt(
            self,
            tr(self.language, "order_qty_title"),
            tr(self.language, "order_qty_label"),
            1,
            1,
            100000,
            1,
        )
        if not ok:
            return
        node_type = str(current.data(5, Qt.ItemDataRole.UserRole) or "")
        added_lines = 0
        warnings: list[str] = []
        qty_factor = float(qty_multiplier)
        selected_only_label = tr(self.language, "order_include_selected")
        children_only_label = tr(self.language, "order_include_children")
        both_label = tr(self.language, "order_include_both")
        include_selected = include_choice == selected_only_label
        include_children = include_choice == children_only_label
        include_full_tree = include_choice == both_label

        if node_type == "article":
            article_id = int(current.data(0, Qt.ItemDataRole.UserRole))
            if include_children:
                for idx in range(current.childCount()):
                    added_lines += self.add_tree_children_to_cart(
                        item=current.child(idx),
                        qty_factor=qty_factor,
                        visited={article_id},
                        depth=0,
                        warnings=warnings,
                    )
            elif include_full_tree:
                for idx in range(current.childCount()):
                    added_lines += self.add_tree_full_to_cart(
                        item=current.child(idx),
                        qty_factor=qty_factor,
                        visited={article_id},
                        depth=0,
                        warnings=warnings,
                    )
        else:
            if include_selected:
                added_lines += 1 if self.add_leaf_to_cart(current, qty_factor=qty_factor) else 0
            if include_children:
                added_lines += self.add_tree_children_to_cart(
                    item=current,
                    qty_factor=qty_factor,
                    visited=set(),
                    depth=0,
                    warnings=warnings,
                    include_self_leaf=False,
                )
            elif include_full_tree:
                added_lines += self.add_tree_full_to_cart(
                    item=current,
                    qty_factor=qty_factor,
                    visited=set(),
                    depth=0,
                    warnings=warnings,
                )
        self.refresh_order_table()
        warn_text = ""
        if warnings:
            unique_warnings = list(dict.fromkeys(warnings))
            warn_text = tr(self.language, "warn_section", items="\n- ".join(unique_warnings))
        QMessageBox.information(
            self,
            tr(self.language, "order_added_title"),
            tr(self.language, "order_added_msg", count=added_lines, warnings=warn_text),
        )

    def add_tree_children_to_cart(
        self,
        item: QTreeWidgetItem,
        qty_factor: float,
        visited: set[int],
        depth: int,
        warnings: list[str],
        include_self_leaf: bool = True,
    ) -> int:
        if depth > 20:
            warnings.append(tr(self.language, "warn_max_depth"))
            return 0
        node_type = str(item.data(5, Qt.ItemDataRole.UserRole) or "")
        line_qty = float(item.data(4, Qt.ItemDataRole.UserRole) or 1.0)
        # Any expandable tree node acts as a parent-only quantity multiplier.
        # It must not be added as a leaf line in children mode.
        if item.childCount() > 0:
            added = 0
            for idx in range(item.childCount()):
                added += self.add_tree_children_to_cart(
                    item=item.child(idx),
                    qty_factor=qty_factor * line_qty,
                    visited=visited,
                    depth=depth + 1,
                    warnings=warnings,
                )
            return added
        if node_type == "assembly_ref":
            child_article_id = item.data(6, Qt.ItemDataRole.UserRole)
            if child_article_id is None:
                child_article_id = self.resolve_article_ref_id(item.text(1))
                if child_article_id is not None:
                    item.setData(6, Qt.ItemDataRole.UserRole, int(child_article_id))
            if child_article_id is None:
                warnings.append(tr(self.language, "warn_missing_subassembly", part=item.text(1)))
                return 0
            return self.add_article_children_to_cart(
                article_id=int(child_article_id),
                qty_factor=qty_factor * line_qty,
                visited=visited | {int(child_article_id)},
                depth=depth + 1,
                warnings=warnings,
            )
        if node_type == "part":
            resolved_article_id = self.resolve_article_ref_id(item.text(1))
            if resolved_article_id is not None:
                item.setData(5, Qt.ItemDataRole.UserRole, "assembly_ref")
                item.setData(6, Qt.ItemDataRole.UserRole, int(resolved_article_id))
                return self.add_article_children_to_cart(
                    article_id=int(resolved_article_id),
                    qty_factor=qty_factor * line_qty,
                    visited=visited | {int(resolved_article_id)},
                    depth=depth + 1,
                    warnings=warnings,
                )
        if node_type == "article":
            article_id = int(item.data(0, Qt.ItemDataRole.UserRole))
            return self.add_article_children_to_cart(
                article_id=article_id,
                qty_factor=qty_factor,
                visited=visited | {article_id},
                depth=depth + 1,
                warnings=warnings,
            )
        if not include_self_leaf:
            return 0
        return 1 if self.add_leaf_to_cart(item, qty_factor=qty_factor) else 0

    def add_tree_full_to_cart(
        self,
        item: QTreeWidgetItem,
        qty_factor: float,
        visited: set[int],
        depth: int,
        warnings: list[str],
    ) -> int:
        if depth > 20:
            warnings.append(tr(self.language, "warn_max_depth"))
            return 0
        node_type = str(item.data(5, Qt.ItemDataRole.UserRole) or "")
        line_qty = float(item.data(4, Qt.ItemDataRole.UserRole) or 1.0)
        added = 0

        if node_type != "article":
            added += 1 if self.add_leaf_to_cart(item, qty_factor=qty_factor) else 0

        if item.childCount() > 0:
            for idx in range(item.childCount()):
                added += self.add_tree_full_to_cart(
                    item=item.child(idx),
                    qty_factor=qty_factor * line_qty,
                    visited=visited,
                    depth=depth + 1,
                    warnings=warnings,
                )
            return added

        if node_type == "assembly_ref":
            child_article_id = item.data(6, Qt.ItemDataRole.UserRole)
            if child_article_id is None:
                child_article_id = self.resolve_article_ref_id(item.text(1))
                if child_article_id is not None:
                    item.setData(6, Qt.ItemDataRole.UserRole, int(child_article_id))
            if child_article_id is None:
                warnings.append(tr(self.language, "warn_missing_subassembly", part=item.text(1)))
                return added
            return added + self.add_article_children_to_cart(
                article_id=int(child_article_id),
                qty_factor=qty_factor * line_qty,
                visited=visited | {int(child_article_id)},
                depth=depth + 1,
                warnings=warnings,
            )

        if node_type == "part":
            resolved_article_id = self.resolve_article_ref_id(item.text(1))
            if resolved_article_id is not None:
                item.setData(5, Qt.ItemDataRole.UserRole, "assembly_ref")
                item.setData(6, Qt.ItemDataRole.UserRole, int(resolved_article_id))
                return added + self.add_article_children_to_cart(
                    article_id=int(resolved_article_id),
                    qty_factor=qty_factor * line_qty,
                    visited=visited | {int(resolved_article_id)},
                    depth=depth + 1,
                    warnings=warnings,
                )

        return added

    def add_article_children_to_cart(
        self,
        article_id: int,
        qty_factor: float,
        visited: set[int],
        depth: int,
        warnings: list[str],
    ) -> int:
        if depth > 20:
            warnings.append(tr(self.language, "warn_max_depth"))
            return 0
        lines = get_article_bom_lines(self.conn, article_id)
        if not lines:
            warnings.append(tr(self.language, "warn_subassembly_no_lines", article_id=article_id))
            return 0
        added = 0
        for line in lines:
            line_part = str(line["part_number"] or "").strip()
            if not line_part:
                continue
            line_qty = float(line["qty"]) if line["qty"] is not None else 1.0
            effective_qty = qty_factor * line_qty
            child_id = self.resolve_article_ref_id(line_part)
            if child_id is not None:
                if child_id in visited:
                    warnings.append(tr(self.language, "warn_cycle_detected", part=line_part))
                    continue
                added += self.add_article_children_to_cart(
                    article_id=child_id,
                    qty_factor=effective_qty,
                    visited=visited | {child_id},
                    depth=depth + 1,
                    warnings=warnings,
                )
                continue
            added += 1 if self.add_leaf_row_to_cart(line, source_article_id=article_id, qty_effective=effective_qty) else 0
        return added

    def add_leaf_to_cart(self, item: QTreeWidgetItem, qty_factor: float) -> bool:
        part_number = str(item.data(1, Qt.ItemDataRole.UserRole) or "").strip()
        if not part_number:
            return False
        part_id = item.data(3, Qt.ItemDataRole.UserRole)
        if part_id is None:
            return False
        article_id = int(self.current_article_id) if self.current_article_id is not None else 0
        item_no = str(item.text(0) or "").strip()
        revision = str(item.data(2, Qt.ItemDataRole.UserRole) or "").strip().upper()
        qty = float(item.data(4, Qt.ItemDataRole.UserRole) or 1.0) * float(qty_factor)
        key = f"{article_id}|{item_no}|{part_number}|{revision}"
        if key not in self.order_cart:
            docs = get_documents_for_part_revision(self.conn, int(part_id), revision or None)
            self.order_cart[key] = {
                "cart_key": key,
                "article_id": article_id,
                "item_no": item_no,
                "part_id": int(part_id),
                "part_number": part_number,
                "revision": revision,
                "description": item.text(3),
                "material": item.text(4),
                "finish": item.text(5),
                "qty_total": 0.0,
                "docs": [str(doc["path"]) for doc in docs if doc["path"]],
                "docs_count": len(docs),
            }
        self.order_cart[key]["qty_total"] = float(self.order_cart[key]["qty_total"]) + qty
        return True

    def add_leaf_row_to_cart(self, line: sqlite3.Row, source_article_id: int, qty_effective: float) -> bool:
        part_number = str(line["part_number"] or "").strip()
        if not part_number:
            return False
        part_id = int(line["part_id"])
        revision = str(line["revision"] or "").strip().upper()
        item_no = str(line["item_no"] or "").strip()
        key = f"{source_article_id}|{item_no}|{part_number}|{revision}"
        if key not in self.order_cart:
            docs = get_documents_for_part_revision(self.conn, part_id, revision or None)
            self.order_cart[key] = {
                "cart_key": key,
                "article_id": int(source_article_id),
                "item_no": item_no,
                "part_id": part_id,
                "part_number": part_number,
                "revision": revision,
                "description": str(line["description"] or ""),
                "material": str(line["material"] or ""),
                "finish": str(line["finish"] or ""),
                "qty_total": 0.0,
                "docs": [str(doc["path"]) for doc in docs if doc["path"]],
                "docs_count": len(docs),
            }
        self.order_cart[key]["qty_total"] = float(self.order_cart[key]["qty_total"]) + float(qty_effective)
        return True

    def refresh_order_table(self) -> None:
        rows = sorted(
            self.order_cart.values(),
            key=lambda row: (
                int(row.get("article_id", 0)),
                str(row.get("item_no", "")),
                str(row.get("part_number", "")),
                str(row.get("revision", "")),
            ),
        )
        self.order_table.setRowCount(len(rows))
        for row_idx, row in enumerate(rows):
            values = [
                str(row.get("part_number") or ""),
                str(row.get("revision") or ""),
                format_qty(float(row.get("qty_total") or 0.0)),
            ]
            for col_idx, value in enumerate(values):
                item = QTableWidgetItem(value)
                if col_idx == 0:
                    item.setData(Qt.ItemDataRole.UserRole, str(row.get("cart_key") or ""))
                self.order_table.setItem(row_idx, col_idx, item)
            delete_button = QPushButton("🗑")
            delete_button.setToolTip(tr(self.language, "btn_clear"))
            row_key = str(row.get("cart_key") or "")
            delete_button.clicked.connect(lambda _checked=False, k=row_key: self.remove_order_line(k))
            self.order_table.setCellWidget(row_idx, 3, delete_button)

    def toggle_order_drawer(self) -> None:
        self.order_drawer_open = not self.order_drawer_open
        self.refresh_order_table()
        self.animate_order_drawer()

    def close_order_drawer(self) -> None:
        if not self.order_drawer_open:
            return
        self.order_drawer_open = False
        self.animate_order_drawer()

    def layout_order_drawer(self) -> None:
        root = self.centralWidget()
        if root is None:
            return
        self.order_backdrop.setGeometry(0, 0, root.width(), root.height())
        self.order_drawer.setGeometry(self.drawer_rect(self.order_drawer_open))
        if self.order_drawer_open:
            self.order_backdrop.raise_()
            self.order_drawer.raise_()

    def drawer_rect(self, is_open: bool) -> QRect:
        root = self.centralWidget()
        if root is None:
            return QRect(0, 0, int(self.order_drawer_width), 0)
        width = int(self.order_drawer_width)
        x = root.width() - width if is_open else root.width()
        return QRect(x, 0, width, root.height())

    def animate_order_drawer(self) -> None:
        root = self.centralWidget()
        if root is None:
            return
        start_rect = self.order_drawer.geometry()
        end_rect = self.drawer_rect(self.order_drawer_open)
        if self.order_backdrop_anim is not None:
            self.order_backdrop_anim.stop()
        if self.order_drawer_anim is not None:
            self.order_drawer_anim.stop()
        target_opacity = 0.45 if self.order_drawer_open else 0.0
        if self.order_drawer_open:
            self.order_backdrop.show()
            self.order_backdrop.raise_()
            if start_rect.width() <= 0 or start_rect.height() <= 0:
                start_rect = self.drawer_rect(False)
                self.order_drawer.setGeometry(start_rect)
        self.order_drawer.raise_()
        self.order_backdrop_anim = QPropertyAnimation(self.order_backdrop, b"opacity", self)
        self.order_backdrop_anim.setDuration(220)
        self.order_backdrop_anim.setEasingCurve(QEasingCurve.Type.InOutCubic)
        self.order_backdrop_anim.setStartValue(self.order_backdrop.get_opacity())
        self.order_backdrop_anim.setEndValue(target_opacity)
        if not self.order_drawer_open:
            def _hide_backdrop_after_close() -> None:
                if not self.order_drawer_open:
                    self.order_backdrop.hide()
            self.order_backdrop_anim.finished.connect(_hide_backdrop_after_close)
        self.order_backdrop_anim.start()
        self.order_drawer_anim = QPropertyAnimation(self.order_drawer, b"geometry", self)
        self.order_drawer_anim.setDuration(240)
        self.order_drawer_anim.setEasingCurve(QEasingCurve.Type.InOutCubic)
        self.order_drawer_anim.setStartValue(start_rect)
        self.order_drawer_anim.setEndValue(end_rect)
        self.order_drawer_anim.start()

    def clear_order_cart(self) -> None:
        if not self.order_cart:
            return
        answer = QMessageBox.question(
            self,
            tr(self.language, "order_clear_title"),
            tr(self.language, "order_clear_msg"),
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if answer != QMessageBox.StandardButton.Yes:
            return
        self.order_cart.clear()
        self.refresh_order_table()

    def remove_order_line(self, cart_key: str) -> None:
        row = self.order_cart.get(cart_key)
        if row is None:
            return
        article_id = int(row.get("article_id") or 0)
        item_no = str(row.get("item_no") or "").strip()
        keys_to_remove = [cart_key]
        child_keys = [
            key
            for key, candidate in self.order_cart.items()
            if key != cart_key
            and int(candidate.get("article_id") or 0) == article_id
            and item_no
            and str(candidate.get("item_no") or "").startswith(item_no + ".")
        ]
        if child_keys:
            answer = QMessageBox.question(
                self,
                tr(self.language, "order_remove_parent_title"),
                tr(self.language, "order_remove_parent_msg", count=len(child_keys)),
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.Yes,
            )
            if answer == QMessageBox.StandardButton.Yes:
                keys_to_remove.extend(child_keys)
        for key in keys_to_remove:
            self.order_cart.pop(key, None)
        self.refresh_order_table()

    def export_order_bundle(self) -> None:
        if not self.order_cart:
            QMessageBox.information(self, tr(self.language, "order_title"), tr(self.language, "order_export_empty"))
            return
        target_root = QFileDialog.getExistingDirectory(self, tr(self.language, "order_export_select_folder"))
        if not target_root:
            return
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        bundle_name = f"ADM_Order_{timestamp}"
        bundle_dir = Path(target_root) / bundle_name
        docs_dir = bundle_dir / "docs"
        docs_dir.mkdir(parents=True, exist_ok=True)

        xlsx_path = bundle_dir / "order_lines.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Order"
        ws.append(["Part NR", "Rev", "Description", "Qty", "Material", "Finish"])
        for row in sorted(
            self.order_cart.values(),
            key=lambda r: (str(r.get("part_number", "")), str(r.get("revision", ""))),
        ):
            ws.append(
                [
                    str(row.get("part_number") or ""),
                    str(row.get("revision") or ""),
                    str(row.get("description") or ""),
                    float(row.get("qty_total") or 0.0),
                    str(row.get("material") or ""),
                    str(row.get("finish") or ""),
                ]
            )
        wb.save(xlsx_path)

        used_names: dict[str, int] = {}
        copied_docs = 0
        for row in self.order_cart.values():
            docs = row.get("docs") or []
            if not isinstance(docs, list):
                continue
            for doc_path in docs:
                src = Path(str(doc_path))
                if not src.exists() or not src.is_file():
                    continue
                base_name = src.name
                if base_name in used_names:
                    used_names[base_name] += 1
                    stem = src.stem
                    suffix = src.suffix
                    out_name = f"{stem}_{used_names[base_name]}{suffix}"
                else:
                    used_names[base_name] = 1
                    out_name = base_name
                shutil.copy2(src, docs_dir / out_name)
                copied_docs += 1

        zip_path = Path(target_root) / f"{bundle_name}.zip"
        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for file in bundle_dir.rglob("*"):
                if file.is_file():
                    zf.write(file, file.relative_to(bundle_dir))

        QMessageBox.information(
            self,
            tr(self.language, "order_export_done_title"),
            tr(self.language, "order_export_done_msg", xlsx=xlsx_path, docs=copied_docs, zip_path=zip_path),
        )

    def on_bom_item_changed(self, item: QTreeWidgetItem, _column: int) -> None:
        if self._suspend_bom_item_changed or not self.developer_mode:
            return
        node_type = str(item.data(5, Qt.ItemDataRole.UserRole) or "")
        if node_type == "article":
            return
        line_id = item.data(0, Qt.ItemDataRole.UserRole)
        if line_id is None:
            return
        self._pending_bom_edits[int(line_id)] = {
            "line_id": int(line_id),
            "item_no": str(item.text(0) or "").strip(),
            "part_number": str(item.text(1) or "").strip(),
            "revision": str(item.text(2) or "").strip(),
            "description": str(item.text(3) or "").strip(),
            "material": str(item.text(4) or "").strip(),
            "finish": str(item.text(5) or "").strip(),
            "qty": str(item.text(6) or "").strip(),
            "line_type": str(item.text(7) or "").strip(),
            "status": str(item.text(8) or "").strip(),
            "source_sheet": str(item.data(7, Qt.ItemDataRole.UserRole) or ""),
            "source_row_number": int(item.data(8, Qt.ItemDataRole.UserRole) or 0),
        }
        self._bom_dirty = True
        self.save_bom_button.setVisible(True)

    def bom_field_for_column(self, col: int) -> str | None:
        mapping = {
            0: "line_no",
            1: "part_number",
            2: "revision",
            3: "description",
            4: "material",
            5: "finish",
            6: "qty",
            7: "line_type",
            8: "status",
        }
        return mapping.get(col)

    def save_bom_edits(self) -> None:
        if not self.developer_mode or not self._pending_bom_edits:
            return
        source_path = Path(self.current_article_source_bom_path)
        if not source_path.exists():
            QMessageBox.warning(self, tr(self.language, "save_bom_failed_title"), tr(self.language, "save_bom_missing_file"))
            return
        if source_path.suffix.lower() == ".xls":
            QMessageBox.warning(self, tr(self.language, "save_bom_failed_title"), tr(self.language, "save_bom_xls_unsupported"))
            return
        try:
            workbook = load_workbook(source_path)
            sheet_header_map: dict[str, dict[int, str]] = {}
            for ws in workbook.worksheets:
                rows = [list(row) for row in ws.iter_rows(values_only=True)]
                _, header_map = detect_header(rows)
                sheet_header_map[ws.title] = header_map

            for edit in self._pending_bom_edits.values():
                sheet_name = str(edit.get("source_sheet") or "")
                row_number = int(edit.get("source_row_number") or 0)
                if not sheet_name or row_number <= 0:
                    continue
                ws = workbook[sheet_name]
                header_map = sheet_header_map.get(sheet_name, {})
                inverse_map = {canonical: idx for idx, canonical in header_map.items()}
                field_values = {
                    "line_no": str(edit.get("item_no") or ""),
                    "part_number": str(edit.get("part_number") or ""),
                    "revision": str(edit.get("revision") or ""),
                    "description": str(edit.get("description") or ""),
                    "material": str(edit.get("material") or ""),
                    "finish": str(edit.get("finish") or ""),
                    "qty": str(edit.get("qty") or ""),
                    "line_type": str(edit.get("line_type") or ""),
                    "status": str(edit.get("status") or ""),
                }
                for field, raw_value in field_values.items():
                    col_idx = inverse_map.get(field)
                    if col_idx is None:
                        continue
                    out_value: object = raw_value
                    if field == "qty":
                        parsed = parse_qty(raw_value)
                        out_value = parsed if parsed is not None else raw_value
                    ws.cell(row=row_number, column=col_idx + 1, value=out_value)

            workbook.save(source_path)
        except Exception as exc:
            QMessageBox.warning(self, tr(self.language, "save_bom_failed_title"), str(exc))
            return

        self._pending_bom_edits = {}
        self._bom_dirty = False
        self.save_bom_button.setVisible(False)
        target_article_id = int(self.current_article_id or 0)
        stats = run_index(self.conn, data_root=self._path_to_root())
        self.refresh_articles()
        if target_article_id > 0:
            if self.select_article_in_table(target_article_id):
                self.display_article_by_id(target_article_id)
            else:
                self.display_article_by_id(target_article_id)
        QMessageBox.information(
            self,
            tr(self.language, "save_bom_done_title"),
            tr(
                self.language,
                "save_bom_done_msg",
                boms=stats.boms_parsed,
                lines=stats.lines_imported,
                warnings=stats.warnings_count,
                errors=stats.errors_count,
            ),
        )

    def run_revision_suggestions(self) -> None:
        if not self.developer_mode or self.current_article_id is None:
            return
        lines = get_article_bom_lines(self.conn, int(self.current_article_id))
        rows: list[dict[str, object]] = []
        for line in lines:
            current_rev = str(line["revision"] or "").strip().upper()
            if current_rev:
                continue
            part_number = str(line["part_number"] or "").strip()
            if not part_number:
                continue
            docs = self.conn.execute(
                "SELECT filename FROM documents WHERE UPPER(filename) LIKE ?",
                (f"%{part_number.upper()}%",),
            ).fetchall()
            found_revs: list[str] = []
            for doc in docs:
                parsed = parse_document_part_and_revision(str(doc["filename"] or ""))
                if not parsed:
                    continue
                parsed_part, parsed_rev = parsed
                if parsed_part != part_number.upper() or not parsed_rev:
                    continue
                rev_letter = parsed_rev.strip().upper()[:1]
                if len(rev_letter) == 1 and rev_letter.isalpha():
                    found_revs.append(rev_letter)
            if not found_revs:
                continue
            chosen = sorted(set(found_revs))[-1]
            rows.append(
                {
                    "line_id": int(line["id"]),
                    "item_no": str(line["item_no"] or ""),
                    "part_number": part_number,
                    "current_revision": current_rev,
                    "suggested_revision": chosen,
                }
            )
        if not rows:
            QMessageBox.information(self, tr(self.language, "revision_suggest_title"), tr(self.language, "revision_suggest_none"))
            return
        dlg = RevisionSuggestionDialog(rows, language=self.language, parent=self)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        selected = dlg.selected_rows()
        if not selected:
            return
        for row in selected:
            line_id = int(row["line_id"])
            item = self.find_tree_item_by_line_id(line_id)
            if item is None:
                continue
            self._suspend_bom_item_changed = True
            item.setText(2, str(row["suggested_revision"]))
            self._suspend_bom_item_changed = False
            self.on_bom_item_changed(item, 2)

    def find_tree_item_by_line_id(self, line_id: int) -> QTreeWidgetItem | None:
        def walk(node: QTreeWidgetItem) -> QTreeWidgetItem | None:
            current_line_id = node.data(0, Qt.ItemDataRole.UserRole)
            if current_line_id is not None and int(current_line_id) == line_id:
                return node
            for i in range(node.childCount()):
                found = walk(node.child(i))
                if found is not None:
                    return found
            return None

        for top_idx in range(self.bom_tree.topLevelItemCount()):
            top = self.bom_tree.topLevelItem(top_idx)
            found = walk(top)
            if found is not None:
                return found
        return None

    def open_settings(self) -> None:
        previous_data_root = self.data_root
        dlg = SettingsDialog(
            self.data_root,
            self.theme_mode,
            self.language,
            developer_mode=self.developer_mode,
            developer_toggle_available=self.developer_toggle_available,
            parent=self,
        )
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        new_root = dlg.selected_data_root()
        if not new_root:
            QMessageBox.warning(self, tr(self.language, "settings_invalid_title"), tr(self.language, "settings_invalid_empty"))
            return
        new_path = Path(new_root)
        if not new_path.exists() or not new_path.is_dir():
            QMessageBox.warning(self, tr(self.language, "settings_invalid_title"), tr(self.language, "settings_invalid_not_exists"))
            return
        self.data_root = str(new_path.resolve())
        self.theme_mode = dlg.selected_theme_mode()
        self.language = dlg.selected_language()
        self.developer_mode = dlg.selected_developer_mode() if self.developer_toggle_available else False
        save_settings(
            AppSettings(
                data_root=self.data_root,
                theme_mode=self.theme_mode,
                has_seen_help=self.has_seen_help,
                language=self.language,
                developer_mode=self.developer_mode,
            )
        )
        self.apply_translations()
        self.apply_theme(self.theme_mode)
        if self.developer_mode:
            self.bom_tree.setEditTriggers(
                QAbstractItemView.EditTrigger.DoubleClicked
                | QAbstractItemView.EditTrigger.SelectedClicked
                | QAbstractItemView.EditTrigger.EditKeyPressed
            )
        else:
            self.bom_tree.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
            self._pending_bom_edits = {}
            self._bom_dirty = False
            self.save_bom_button.setVisible(False)
        self.revision_suggest_button.setVisible(self.developer_mode)
        if str(previous_data_root).strip().lower() != str(self.data_root).strip().lower():
            self.reindex()

    def set_preview_message(self, message: str) -> None:
        self.preview_message.setText(message)
        self.preview_stack.setCurrentWidget(self.preview_message)
        self.current_pdf_path = ""
        self.current_pdf_page = 0
        self.current_pdf_page_count = 0
        self.update_preview_controls()

    def update_preview_controls(self) -> None:
        if self.current_pdf_page_count <= 0:
            self.preview_prev_button.setEnabled(False)
            self.preview_next_button.setEnabled(False)
            self.preview_page_label.setText("-")
            self.preview_page_hint.setText("")
            return
        self.preview_prev_button.setEnabled(self.current_pdf_page > 0)
        self.preview_next_button.setEnabled(self.current_pdf_page < self.current_pdf_page_count - 1)
        self.preview_page_label.setText(
            tr(
                self.language,
                "preview_page_label",
                page=self.current_pdf_page + 1,
                count=self.current_pdf_page_count,
            )
        )
        if self.current_pdf_page_count > 1:
            self.preview_page_hint.setText(tr(self.language, "preview_multi_page"))
        else:
            self.preview_page_hint.setText("")

    def preview_prev_page(self) -> None:
        if self.current_pdf_page_count <= 0:
            return
        self.set_pdf_page(self.current_pdf_page - 1)

    def preview_next_page(self) -> None:
        if self.current_pdf_page_count <= 0:
            return
        self.set_pdf_page(self.current_pdf_page + 1)

    def set_pdf_page(self, page: int) -> None:
        if self.pdf_view is None or self.current_pdf_page_count <= 0:
            return
        page_safe = max(0, min(int(page), self.current_pdf_page_count - 1))
        self.current_pdf_page = page_safe
        navigator = self.pdf_view.pageNavigator()
        try:
            navigator.jump(page_safe, QPointF(), 0.0)
        except TypeError:
            navigator.jump(page_safe, QPointF())
        self.update_preview_controls()

    def preview_first_pdf_in_list(self) -> None:
        for idx in range(self.docs_list.count()):
            item = self.docs_list.item(idx)
            path = str(item.data(Qt.ItemDataRole.UserRole) or "")
            if Path(path).suffix.lower() == ".pdf":
                self.docs_list.setCurrentRow(idx)
                self.preview_pdf(path)
                return
        self.set_preview_message(tr(self.language, "preview_no_pdf"))

    def preview_selected_document(self) -> None:
        selected = self.docs_list.selectedItems()
        if not selected:
            return
        path = str(selected[0].data(Qt.ItemDataRole.UserRole) or "")
        if Path(path).suffix.lower() != ".pdf":
            self.set_preview_message(tr(self.language, "preview_not_pdf"))
            return
        self.preview_pdf(path)

    def preview_pdf(self, path: str) -> None:
        if not path:
            self.set_preview_message(tr(self.language, "preview_no_selected"))
            return
        if not PDF_PREVIEW_AVAILABLE or self.pdf_document is None or self.pdf_view is None:
            self.set_preview_message(tr(self.language, "preview_component_missing"))
            return
        load_result = self.pdf_document.load(path)
        no_error = getattr(QPdfDocument.Error, "None_", None)
        if no_error is None:
            no_error = getattr(QPdfDocument.Error, "NoError", None)
        if no_error is not None and load_result != no_error:
            self.set_preview_message(tr(self.language, "preview_failed_load", result=load_result))
            return
        if no_error is None and str(load_result).lower() not in ("error.none", "none", "noerror", "error.noerror"):
            self.set_preview_message(tr(self.language, "preview_failed_load_generic"))
            return
        self.current_pdf_path = path
        self.current_pdf_page_count = int(self.pdf_document.pageCount())
        self.current_pdf_page = 0
        self.pdf_view.verticalScrollBar().setValue(0)
        self.pdf_view.horizontalScrollBar().setValue(0)
        self.set_pdf_page(0)
        self.preview_stack.setCurrentWidget(self.pdf_view)
        self.update_preview_controls()

    def apply_status_style(self, item: QTreeWidgetItem, status_text: str, column: int) -> None:
        normalized = (status_text or "").strip().lower()
        item.setForeground(column, QColor(255, 255, 255))
        if normalized in {"approved", "released"}:
            item.setBackground(column, QColor(30, 120, 55))
        elif normalized == "denied":
            item.setBackground(column, QColor(150, 35, 35))
        else:
            item.setBackground(column, QColor(170, 100, 20))

    def apply_theme(self, theme_mode: str) -> None:
        apply_app_theme(theme_mode)

    def resolve_help_pdf_path(self) -> Path | None:
        return None

    def open_help_manual_clicked(self) -> None:
        self.open_help_manual(mark_seen=True, show_warning=True)

    def open_help_manual_first_run(self) -> None:
        self.open_help_manual(mark_seen=True, show_warning=True)

    def open_help_manual(self, mark_seen: bool, show_warning: bool) -> None:
        _ = show_warning
        dlg = HelpDialog(language=self.language, parent=self)
        dlg.exec()

        if mark_seen and not self.has_seen_help:
            self.has_seen_help = True
            save_settings(
                AppSettings(
                    data_root=self.data_root,
                    theme_mode=self.theme_mode,
                    has_seen_help=True,
                    language=self.language,
                    developer_mode=self.developer_mode,
                )
            )

    def check_for_updates(self) -> None:
        if self._update_check_done:
            return
        self._update_check_done = True
        latest = fetch_latest_github_release()
        if not latest:
            return
        latest_tag = str(latest.get("tag_name") or "").strip()
        if not latest_tag:
            return
        if not is_newer_version(__version__, latest_tag):
            return
        message = (
            tr(self.language, "update_available_msg", installed=__version__, latest=latest_tag)
        )
        answer = QMessageBox.question(
            self,
            tr(self.language, "update_available_title"),
            message,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes,
        )
        if answer == QMessageBox.StandardButton.Yes:
            open_file(RELEASES_URL)


def apply_app_theme(theme_mode: str) -> None:
    app = QApplication.instance()
    if app is None:
        return
    app.setStyle("Fusion")
    mode = theme_mode if theme_mode in {"light", "dark"} else "light"
    if mode == "dark":
        palette = QPalette()
        palette.setColor(QPalette.ColorRole.Window, QColor("#1E1E1E"))
        palette.setColor(QPalette.ColorRole.WindowText, QColor("#EDEDED"))
        palette.setColor(QPalette.ColorRole.Base, QColor("#252525"))
        palette.setColor(QPalette.ColorRole.AlternateBase, QColor("#1E1E1E"))
        palette.setColor(QPalette.ColorRole.ToolTipBase, QColor("#252525"))
        palette.setColor(QPalette.ColorRole.ToolTipText, QColor("#EDEDED"))
        palette.setColor(QPalette.ColorRole.Text, QColor("#F2F2F2"))
        palette.setColor(QPalette.ColorRole.Button, QColor("#2B2B2B"))
        palette.setColor(QPalette.ColorRole.ButtonText, QColor("#F2F2F2"))
        palette.setColor(QPalette.ColorRole.BrightText, QColor(255, 255, 255))
        palette.setColor(QPalette.ColorRole.Highlight, QColor("#97C21E"))
        palette.setColor(QPalette.ColorRole.HighlightedText, QColor(255, 255, 255))
        app.setPalette(palette)
        app.setStyleSheet(
            """
            QLineEdit, QTableWidget, QTreeWidget, QListWidget, QComboBox {
                background: #252525;
                color: #F2F2F2;
                border: 1px solid #3A3A3A;
            }
            QPushButton {
                background: #2B2B2B;
                color: #F2F2F2;
                border: 1px solid #444444;
                padding: 4px 10px;
            }
            QPushButton:hover { background: #333333; }
            QHeaderView::section {
                background: #2F2F2F;
                color: #F2F2F2;
                border: 1px solid #3A3A3A;
                padding: 4px;
            }
            """
        )
        return

    palette = QPalette()
    palette.setColor(QPalette.ColorRole.Window, QColor("#FFFFFF"))
    palette.setColor(QPalette.ColorRole.WindowText, QColor("#111111"))
    palette.setColor(QPalette.ColorRole.Base, QColor("#FFFFFF"))
    palette.setColor(QPalette.ColorRole.AlternateBase, QColor("#F7F7F7"))
    palette.setColor(QPalette.ColorRole.ToolTipBase, QColor("#FFFFFF"))
    palette.setColor(QPalette.ColorRole.ToolTipText, QColor("#111111"))
    palette.setColor(QPalette.ColorRole.Text, QColor("#111111"))
    palette.setColor(QPalette.ColorRole.Button, QColor("#F3F4F6"))
    palette.setColor(QPalette.ColorRole.ButtonText, QColor("#111111"))
    palette.setColor(QPalette.ColorRole.BrightText, QColor("#111111"))
    palette.setColor(QPalette.ColorRole.Highlight, QColor("#97C21E"))
    palette.setColor(QPalette.ColorRole.HighlightedText, QColor("#111111"))
    app.setPalette(palette)
    app.setStyleSheet(
        """
        QLineEdit, QTableWidget, QTreeWidget, QListWidget, QComboBox {
            background: #FFFFFF;
            color: #111111;
            border: 1px solid #D1D5DB;
        }
        QPushButton {
            background: #F3F4F6;
            color: #111111;
            border: 1px solid #D1D5DB;
            padding: 4px 10px;
        }
        QPushButton:hover { background: #E5E7EB; }
        QHeaderView::section {
            background: #F9FAFB;
            color: #111111;
            border: 1px solid #D1D5DB;
            padding: 4px;
        }
        """
    )


def format_qty(value: float) -> str:
    if float(value).is_integer():
        return str(int(value))
    return f"{value:.4f}".rstrip("0").rstrip(".")


def open_file(path: str) -> None:
    if not path:
        return
    if sys.platform.startswith("win"):
        os.startfile(path)  # type: ignore[attr-defined]
        return
    subprocess.run(["xdg-open", path], check=False)


def run_ui(
    conn: sqlite3.Connection,
    data_root: str,
    theme_mode: str = "light",
    has_seen_help: bool = False,
    language: str = "en",
    developer_mode: bool = False,
) -> int:
    app = QApplication.instance() or QApplication(sys.argv)
    icon = get_app_icon()
    if icon is not None:
        app.setWindowIcon(icon)
    apply_app_theme(theme_mode)
    startup_dialog = QDialog()
    startup_dialog.setWindowTitle(tr(language, "startup_reindex_title"))
    startup_dialog.setModal(True)
    startup_dialog.setWindowFlag(Qt.WindowType.WindowCloseButtonHint, False)
    startup_dialog.setWindowFlag(Qt.WindowType.WindowContextHelpButtonHint, False)
    startup_dialog.setFixedSize(460, 130)
    startup_layout = QVBoxLayout(startup_dialog)
    startup_layout.addWidget(QLabel(tr(language, "startup_reindex_msg")))
    startup_bar = QProgressBar()
    startup_bar.setRange(0, 0)
    startup_layout.addWidget(startup_bar)
    startup_status = QLabel(tr(language, "startup_reindex_running"))
    startup_layout.addWidget(startup_status)

    startup_result: dict[str, object] = {"stats": None, "error": None, "done": False}
    db_row = conn.execute("PRAGMA database_list").fetchone()
    db_path = Path(str(db_row["file"] if db_row and "file" in db_row.keys() else "")).resolve()
    data_root_path = Path(data_root).resolve()

    def _index_in_background() -> None:
        from .db import get_connection

        local_conn = get_connection(db_path)
        try:
            startup_result["stats"] = run_index(local_conn, data_root_path)
        except Exception as exc:
            startup_result["error"] = str(exc)
        finally:
            local_conn.close()
            startup_result["done"] = True

    worker = threading.Thread(target=_index_in_background, daemon=True)
    worker.start()

    poll_timer = QTimer(startup_dialog)

    def _poll_status() -> None:
        if not bool(startup_result.get("done")):
            return
        poll_timer.stop()
        if startup_result.get("error"):
            startup_status.setText(tr(language, "startup_reindex_failed"))
            startup_dialog.reject()
            return
        startup_status.setText(tr(language, "startup_reindex_done"))
        startup_dialog.accept()

    poll_timer.setInterval(100)
    poll_timer.timeout.connect(_poll_status)
    poll_timer.start()
    startup_dialog.exec()
    worker.join(timeout=10.0)

    if startup_result["error"]:
        QMessageBox.warning(
            None,
            tr(language, "startup_reindex_error_title"),
            tr(language, "startup_reindex_error_msg", error=str(startup_result["error"])),
        )
    win = MainWindow(
        conn,
        data_root=data_root,
        theme_mode=theme_mode,
        has_seen_help=has_seen_help,
        language=language,
        developer_mode=developer_mode,
    )
    win.show()
    return app.exec()
