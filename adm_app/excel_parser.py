from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Iterator

from openpyxl import load_workbook
import xlrd

from .mapping import map_headers, parse_qty


ARTICLE_PATTERN = re.compile(r"(?i)\bBOM\s*(\d{3,})\b")
FALLBACK_ARTICLE_PATTERN = re.compile(r"\b(\d{3,})\b")


@dataclass
class ParsedLine:
    part_number: str
    description: str | None
    qty: float | None
    revision: str | None
    material: str | None
    finish: str | None
    line_type: str | None
    status: str | None
    unit: str | None
    item_no: str | None
    line_no: int | None
    raw_columns: dict[str, object]
    source_sheet: str
    source_row_number: int


@dataclass
class ParsedBOM:
    article_number: str
    article_title: str
    lines: list[ParsedLine]
    source_file: Path


def parse_bom_file(path: Path) -> ParsedBOM:
    article_number, article_title = extract_article_from_filename(path.name)
    lines: list[ParsedLine] = []
    for sheet_name, rows in iter_sheet_rows(path):
        lines.extend(parse_sheet_rows(sheet_name, rows))
    return ParsedBOM(
        article_number=article_number,
        article_title=article_title,
        lines=lines,
        source_file=path,
    )


def extract_article_from_filename(filename: str) -> tuple[str, str]:
    stem = Path(filename).stem
    match = ARTICLE_PATTERN.search(stem) or FALLBACK_ARTICLE_PATTERN.search(stem)
    article = match.group(1) if match else f"UNKNOWN_{stem}"
    title = re.sub(r"(?i)^bom\s*", "", stem).strip()
    title = re.sub(rf"^{re.escape(article)}\s*", "", title).strip()
    return article, title or stem


def iter_sheet_rows(path: Path) -> Iterator[tuple[str, list[list[object]]]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        wb = load_workbook(path, data_only=True, read_only=True)
        for ws in wb.worksheets:
            rows = [list(row) for row in ws.iter_rows(values_only=True)]
            yield ws.title, rows
        return
    if suffix == ".xls":
        book = xlrd.open_workbook(path.as_posix())
        for sheet in book.sheets():
            rows = [sheet.row_values(i) for i in range(sheet.nrows)]
            yield sheet.name, rows
        return
    raise ValueError(f"Unsupported extension: {path.suffix}")


def parse_sheet_rows(sheet_name: str, rows: list[list[object]]) -> list[ParsedLine]:
    header_idx, header_map = detect_header(rows)
    if header_idx is None or "part_number" not in header_map.values():
        return []
    headers = rows[header_idx]
    parsed: list[ParsedLine] = []
    for row_number, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        part = read_cell(row, header_map, "part_number")
        if part is None:
            continue
        raw_columns: dict[str, object] = {}
        for idx, value in enumerate(row):
            if idx in header_map:
                continue
            key = str(headers[idx]).strip() if idx < len(headers) else f"col_{idx + 1}"
            if value not in (None, ""):
                raw_columns[key] = value
        parsed.append(
            ParsedLine(
                part_number=str(part).strip(),
                description=to_text(read_cell(row, header_map, "description")),
                qty=parse_qty(read_cell(row, header_map, "qty")),
                revision=to_text(read_cell(row, header_map, "revision")),
                material=to_text(read_cell(row, header_map, "material")),
                finish=to_text(read_cell(row, header_map, "finish")),
                line_type=to_text(read_cell(row, header_map, "line_type")),
                status=to_text(read_cell(row, header_map, "status")),
                unit=to_text(read_cell(row, header_map, "unit")),
                item_no=to_item_no(read_cell(row, header_map, "line_no")),
                line_no=to_int(read_cell(row, header_map, "line_no")),
                raw_columns=raw_columns,
                source_sheet=sheet_name,
                source_row_number=row_number,
            )
        )
    return parsed


def detect_header(rows: list[list[object]], max_scan: int = 30) -> tuple[int | None, dict[int, str]]:
    best_index: int | None = None
    best_map: dict[int, str] = {}
    best_score = -1
    for idx, row in enumerate(rows[:max_scan]):
        mapping = map_headers(row)
        score = len(mapping)
        if score > best_score and "part_number" in mapping.values():
            best_index = idx
            best_map = mapping
            best_score = score
    return best_index, best_map


def read_cell(row: list[object], mapping: dict[int, str], field: str) -> object | None:
    for idx, canonical in mapping.items():
        if canonical == field and idx < len(row):
            value = row[idx]
            if isinstance(value, str) and not value.strip():
                return None
            return value
    return None


def to_text(value: object | None) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def to_int(value: object | None) -> int | None:
    if value is None:
        return None
    if isinstance(value, float):
        return int(value) if value.is_integer() else None
    if isinstance(value, int):
        return value
    text = str(value).strip()
    if not text:
        return None
    if "." in text and not text.endswith(".0"):
        return None
    try:
        parsed = float(text)
        if not parsed.is_integer():
            return None
        return int(parsed)
    except (ValueError, TypeError):
        return None


def to_item_no(value: object | None) -> str | None:
    if value is None:
        return None
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        text = f"{value:.12g}"
        return text.rstrip(".")
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(" ", "")
    text = re.sub(r"\.+", ".", text).strip(".")
    return text or None
